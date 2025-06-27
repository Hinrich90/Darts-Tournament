#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed May 7 20:39:48 2025
@author: hinrich
"""

#------------------------------------------------------------------------------
# Dateiname: Darts.py
# Version: 0.3
# Funktion: Auswertung eines Dartturniers im Highscore- und Rundenmodus
# Autor: Hinrich Gruß
# Datum der letzten Änderung: 21.06.2025
# Änderung zur Vorgaengerversion:
# - Vorbereitung für Android
#------------------------------------------------------------------------------


import sys
import os
import logging
import re
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QComboBox, QLineEdit, QPushButton, QGridLayout, QMessageBox,
                             QTableWidget, QTableWidgetItem, QProgressBar, QFileDialog)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap, QIntValidator
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

import resources  # Importiere die Ressourcendatei

# Umgebungsvariablen fuer Unicode und Qt-Warnungen
os.environ['PYTHONIOENCODING'] = 'utf-8'
os.environ['QT_LOGGING_RULES'] = 'qt5ct.debug=false'

# Logging fuer Debugging
logging.basicConfig(filename='darts.log', level=logging.DEBUG, encoding='utf-8')


const_defaultHighscore = 10000
const_anzeigeRunden = 8
const_groesseScheibe = 480

class DartscheibeLabel(QLabel):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Aktiviere Mausverfolgung, um Bewegungen ohne Klick zu erfassen
        self.setMouseTracking(True)
        # Referenz auf das SpielGUI-Objekt für Punktzahl-Berechnung
        self.spiel_gui = None
        self.original_pixmap = None  # Speichert die ursprüngliche Pixmap
        self.setAlignment(Qt.AlignCenter)
        self.setScaledContents(False)  # Verhindert automatische Skalierung mit Verzerrung

    def set_spiel_gui(self, spiel_gui):
        # Setze die Referenz auf das SpielGUI-Objekt
        self.spiel_gui = spiel_gui

    def set_pixmap(self, pixmap):
        self.original_pixmap = pixmap
        self.update_pixmap()

    def update_pixmap(self):
        if not self.original_pixmap or self.original_pixmap.isNull():
            return
        # Skaliere die Pixmap basierend auf der aktuellen Label-Größe, behalte Seitenverhältnis bei
        label_size = min(self.width(), self.height())
        scaled_pixmap = self.original_pixmap.scaled(
            label_size, label_size,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        super().setPixmap(scaled_pixmap)

    def resizeEvent(self, event):
        # Bei Größenänderung des Labels Pixmap neu skalieren
        self.update_pixmap()
        super().resizeEvent(event)

    def mouseMoveEvent(self, event):
        # Behandle Mausbewegungen über dem Label
        if self.spiel_gui and self.pixmap() and not self.pixmap().isNull():
            # Hole die Mausposition relativ zum Label
            x = event.x()
            y = event.y()
            # Berechne die Punktzahl basierend auf den Koordinaten
            punktzahl = self.spiel_gui.AuswertenScheibe(x, y, self.width(), self.height())
            # Aktualisiere das Punktzahl-Label im SpielGUI
            self.spiel_gui.update_punktzahl_label(punktzahl)
        super().mouseMoveEvent(event)
        
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self.spiel_gui and self.pixmap():
            x = event.x()
            y = event.y()
            punktzahl = self.spiel_gui.AuswertenScheibe(x, y, self.width(), self.height())
            self.spiel_gui.verarbeite_wurf(punktzahl)
        super().mousePressEvent(event)

class Spieler:
    def __init__(self, vorname, nachname, startnr):
        self.vorname = vorname
        self.nachname = nachname
        self.startnr = startnr

class SetupLogik:
    def __init__(self):
        self.spielerliste = []
        self.anzZeilen = 1

    def schaetzeZeit(self, modus, highscore_runden, anz_spieler):
        try:
            hr = int(highscore_runden)
            if hr <= 0:
                logging.warning("Highscore/Rundenanzahl <= 0")
                return [0, 0]
            if modus == "Highscore":
                t = (hr * anz_spieler) / 5185.92 * 60
            else:
                t = (hr * anz_spieler) / 111 * 60
            t_h = int(t // 60)
            t_m = int(t % 60)
            return [t_h, t_m]
        except ValueError:
            logging.warning(f"Ungültige Eingabe fuer Highscore/Runden: {highscore_runden}")
            return [0, 0]

    def import_spieler_excel(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            spieler_daten = []
            
            # Prüfe Kopfzeile, um die Spalten "Vorname" und "Nachname" zu finden
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            vorname_col = None
            nachname_col = None
            for idx, header in enumerate(header_row):
                if header == "Vorname":
                    vorname_col = idx
                elif header == "Nachname":
                    nachname_col = idx
            
            if vorname_col is None or nachname_col is None:
                raise ValueError("Excel-Datei enthält nicht die erforderlichen Spalten 'Vorname' und 'Nachname'")
            
            # Lies die Spielerdaten ab Zeile 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                vorname = row[vorname_col].strip() if row[vorname_col] else ""
                nachname = row[nachname_col].strip() if row[nachname_col] else ""
                if vorname and nachname:  # Nur hinzufügen, wenn beide Felder gefüllt sind
                    spieler_daten.append((vorname, nachname))
            
            logging.info(f"Excel-Datei erfolgreich gelesen: {len(spieler_daten)} Spieler gefunden")
            return spieler_daten
        except Exception as e:
            logging.error(f"Fehler beim Lesen der Excel-Datei: {str(e)}")
            raise e

    def pruefeNamen(self, namen):
        # Ignoriere leere Zeilen am Ende
        gefilterte_namen = [(vorname, nachname) for vorname, nachname in namen if vorname or nachname]
        
        # Pruefe, ob es teilweise gefuellte Zeilen gibt
        for vorname, nachname in gefilterte_namen:
            if (vorname and not nachname) or (not vorname and nachname):
                return False
            # Validiere Namen: Nur Buchstaben, Umlaute, Bindestriche, Leerzeichen
            if vorname and not re.match(r'^[a-zA-Zaeoeueaeoeueß\-\s]+$', vorname):
                return False
            if nachname and not re.match(r'^[a-zA-Zaeoeueaeoeueß\-\s]+$', nachname):
                return False
        
        # Pruefe, ob es mindestens eine gefuellte Zeile gibt
        gefuellte_zeilen = sum(1 for vorname, nachname in gefilterte_namen if vorname and nachname)
        return gefuellte_zeilen >= 1

class SetupWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.logik = SetupLogik()
        self.setWindowTitle("Einstellungen")
        self.const_Zeilenhoehe = 35 if not self.is_android() else 60
        self.const_Abstand = 50 if not self.is_android() else 80
        self.listeVornamenFelder = []
        self.listeNachnamenFelder = []
        self.anzZeilen = self.logik.anzZeilen
        self.anzahl_spieler = 1
        self.debounce_timer = QTimer()
        self.debounce_timer.setSingleShot(True)
        self.debounce_timer.timeout.connect(self._pruefeNamen_debounced)
        self.initUI()

    def is_android(self):
        import platform
        return platform.system().lower() == 'android' or 'android' in sys.platform.lower()

    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Stylesheet fuer touch-freundliche GUI
        if self.is_android():
            self.setStyleSheet("""
                QLabel { font-size: 18pt; }
                QLineEdit { font-size: 18pt; padding: 12px; min-height: 50px; }
                QComboBox { font-size: 18pt; padding: 12px; min-height: 50px; }
                QPushButton { font-size: 18pt; padding: 15px; min-height: 60px; min-width: 150px; }
                QPushButton:pressed { background-color: #cccccc; }
            """)
        else:
            self.setStyleSheet("""
                QLabel { font-size: 12pt; }
                QLineEdit { font-size: 12pt; padding: 4px; }
                QComboBox { font-size: 12pt; padding: 4px; }
                QPushButton { font-size: 12pt; padding: 6px; }
            """)

        modus_layout = QHBoxLayout()
        self.lbl_modus = QLabel("Spielmodus:")
        self.combo_modus = QComboBox()
        self.combo_modus.addItems(["Highscore", "Rundenwertung"])
        self.combo_modus.currentTextChanged.connect(self.update_highscoreRunden_label)
        modus_layout.addWidget(self.lbl_modus)
        modus_layout.addWidget(self.combo_modus)
        main_layout.addLayout(modus_layout)

        highscore_layout = QHBoxLayout()
        self.lbl_highscoreRunden = QLabel("Highscore:")
        self.ent_highscoreRunden = QLineEdit(str(const_defaultHighscore))
        self.ent_highscoreRunden.setValidator(QIntValidator(1, 999999))                         # nur positive Ganzzahlen zulassen
        self.ent_highscoreRunden.textChanged.connect(self.update_schaetzung)
        self.ent_highscoreRunden.textChanged.connect(self.update_highscoreRunden_label)
        highscore_layout.addWidget(self.lbl_highscoreRunden)
        highscore_layout.addWidget(self.ent_highscoreRunden)
        main_layout.addLayout(highscore_layout)

        passwort_layout = QHBoxLayout()
        self.lbl_passwort = QLabel("Passwort festlegen:")
        self.ent_passwort = QLineEdit()
        self.ent_passwort.setEchoMode(QLineEdit.Password)
        passwort_layout.addWidget(self.lbl_passwort)
        passwort_layout.addWidget(self.ent_passwort)
        main_layout.addLayout(passwort_layout)

        passwort2_layout = QHBoxLayout()
        self.lbl_passwort2 = QLabel("Passwort wiederholen:")
        self.ent_passwort2 = QLineEdit()
        self.ent_passwort2.setEchoMode(QLineEdit.Password)
        passwort2_layout.addWidget(self.lbl_passwort2)
        passwort2_layout.addWidget(self.ent_passwort2)
        main_layout.addLayout(passwort2_layout)

        self.lbl_erklaerung = QLabel(f"Das Spiel endet, wenn ein Spieler {const_defaultHighscore} Punkte erreicht hat.")
        self.lbl_erklaerung.setWordWrap(True)
        main_layout.addWidget(self.lbl_erklaerung)

        self.lbl_schaetzung = QLabel("Geschätzte Spielzeit: 0 h, 0 min")
        main_layout.addWidget(self.lbl_schaetzung)

        button_layout = QHBoxLayout()
        self.btn_start = QPushButton("Start")
        self.btn_start.clicked.connect(self.start)
        self.btn_import = QPushButton("Import aus Turnier")
        self.btn_import.clicked.connect(self.importTurnier)
        button_layout.addWidget(self.btn_start)
        button_layout.addWidget(self.btn_import)
        main_layout.addLayout(button_layout)

        self.spieler_grid = QGridLayout()
        self.lbl_vornamen = QLabel("Vorname")
        self.lbl_nachnamen = QLabel("Nachname")
        self.spieler_grid.addWidget(self.lbl_vornamen, 0, 0)
        self.spieler_grid.addWidget(self.lbl_nachnamen, 0, 1)
        for i in range(self.anzZeilen):
            self.ergaenzeZeile(i + 1)
        main_layout.addLayout(self.spieler_grid)

        self.update_schaetzung()

    def ergaenzeZeile(self, zeile):
        ent_vorname = QLineEdit()
        ent_nachname = QLineEdit()
        ent_vorname.textChanged.connect(self.pruefeNamen)
        ent_nachname.textChanged.connect(self.pruefeNamen)
        self.spieler_grid.addWidget(ent_vorname, zeile, 0)
        self.spieler_grid.addWidget(ent_nachname, zeile, 1)
        self.listeVornamenFelder.append(ent_vorname)
        self.listeNachnamenFelder.append(ent_nachname)

    def pruefeNamen(self):
        self.debounce_timer.start(200)

    def _pruefeNamen_debounced(self):
        namen = [(f.text().strip(), n.text().strip()) for f, n in zip(self.listeVornamenFelder, self.listeNachnamenFelder)]
        
        # Prüfe, ob die Namen gültig sind
        if not self.logik.pruefeNamen(namen):
            for vorname, nachname in namen:
                if vorname and not re.match(r'^[a-zA-ZäöüÄÖÜß\-\s]+$', vorname):
                    QMessageBox.warning(self, "Fehler", f"Ungültiger Vorname: {vorname}")
                    return
                if nachname and not re.match(r'^[a-zA-ZäöüÄÖÜß\-\s]+$', nachname):
                    QMessageBox.warning(self, "Fehler", f"Ungültiger Nachname: {nachname}")
                    return
            return
        self.anzahl_spieler = sum(1 for vorname, nachname in namen if vorname and nachname)
        self.update_schaetzung()
        # Finde die letzte gefüllte Zeile
        letzte_gefuellte_zeile = -1
        for i, (vorname, nachname) in enumerate(namen):
            if vorname and nachname:
                letzte_gefuellte_zeile = i

        # Füge eine neue Zeile hinzu, nur wenn die letzte Zeile gefüllt ist
        # (d.h. keine leere Zeile mehr am Ende)
        if letzte_gefuellte_zeile == len(namen) - 1 and len(self.listeVornamenFelder) == self.anzZeilen:
            self.anzZeilen += 1
            self.ergaenzeZeile(self.anzZeilen)
            logging.debug(f"Neue Zeile hinzugefügt: {self.anzZeilen}")
        #self.update_schaetzung()

    def update_highscoreRunden_label(self):
        modus = self.combo_modus.currentText()
        if modus == "Highscore":
            self.lbl_highscoreRunden.setText("Highscore:")
            self.lbl_erklaerung.setText(f"Das Spiel endet, wenn ein Spieler {self.ent_highscoreRunden.text().strip()} Punkte erreicht hat.")
        else:
            self.lbl_highscoreRunden.setText("Anzahl Runden:")
            self.lbl_erklaerung.setText(f"Das Spiel endet nach {self.ent_highscoreRunden.text().strip()} Runden.")
        self.update_schaetzung()

    def update_schaetzung(self):
        t = self.logik.schaetzeZeit(self.combo_modus.currentText(), self.ent_highscoreRunden.text().strip(), self.anzahl_spieler)
        self.lbl_schaetzung.setText(f"Geschätzte Spielzeit: {t[0]} h, {t[1]} min")

    def importTurnier(self):
        if self.is_android():
            try:
                from jnius import autoclass
                Environment = autoclass('android.os.Environment')
                Context = autoclass('org.kivy.android.PythonActivity').mActivity
                storage_dir = Context.getExternalFilesDir(None).getAbsolutePath()
                file_path = os.path.join(storage_dir, "Turnierergebnisse.xlsx")
                logging.info(f"Android-Dateipfad: {file_path}")
            except ImportError:
                logging.error("Pyjnius nicht verfügbar, fallback auf Platzhalter")
                file_path = "/sdcard/Darts/Turnierergebnisse.xlsx"
        else:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Wähle eine Excel-Datei aus",
                "",
                "Excel-Dateien (*.xlsx);;Alle Dateien (*.*)"
            )
        if not file_path:
            return

        try:
            spieler_daten = self.logik.import_spieler_excel(file_path)

            if not spieler_daten:
                QMessageBox.information(self, "Fehler", "Die ausgewählte Excel-Datei enthält keine gültigen Spielerdaten.")
                return

            # Finde die erste leere Zeile (oder die erste Zeile, die nur teilweise gefüllt ist)
            bestehende_namen = [
                (f.text().strip(), n.text().strip())
                for f, n in zip(self.listeVornamenFelder, self.listeNachnamenFelder)
            ]
            erste_leere_zeile = 0
            for i, (vorname, nachname) in enumerate(bestehende_namen):
                if not vorname and not nachname:
                    erste_leere_zeile = i
                    break
                elif (vorname and not nachname) or (not vorname and nachname):
                    QMessageBox.warning(self, "Fehler", f"Zeile {i+1}: Vor- und Nachname müssen beide angegeben sein.")
                    return
                erste_leere_zeile = i + 1

            # Füge neue Zeilen hinzu, falls nötig
            benötigte_zeilen = erste_leere_zeile + len(spieler_daten) + 1 - len(self.listeVornamenFelder)
            for _ in range(max(0, benötigte_zeilen)):
                self.anzZeilen += 1
                self.ergaenzeZeile(self.anzZeilen)

            # Trage importierte Namen ein
            for i, (vorname, nachname) in enumerate(spieler_daten):
                zeile = erste_leere_zeile + i
                self.listeVornamenFelder[zeile].blockSignals(True)
                self.listeNachnamenFelder[zeile].blockSignals(True)
                self.listeVornamenFelder[zeile].setText(vorname)
                self.listeNachnamenFelder[zeile].setText(nachname)
                self.listeVornamenFelder[zeile].blockSignals(False)
                self.listeNachnamenFelder[zeile].blockSignals(False)

            QMessageBox.information(self, "Erfolg", f"{len(spieler_daten)} Spieler erfolgreich importiert.")
            self.update_schaetzung()
        except Exception as e:
            error_msg = str(e).encode('ascii', 'replace').decode('ascii')
            QMessageBox.critical(self, "Fehler", f"Fehler beim Importieren der Excel-Datei: {error_msg}")

    def start(self):
        if self.ent_passwort.text() != self.ent_passwort2.text():
            QMessageBox.critical(self, "Fehler", "Passwörter stimmen nicht überein!")
            self.ent_passwort.clear()
            self.ent_passwort2.clear()
            return
        
        # Validiere Highscore/Runden-Eingabe
        try:
            hr = int(self.ent_highscoreRunden.text().strip())
            if hr <= 0:
                raise ValueError("Highscore/Anzahl Runden muss größer als 0 sein.")
        except ValueError:
            QMessageBox.critical(self, "Fehler", "Bitte geben Sie eine gültige Zahl für Highscore/Anzahl Runden ein.")
            return
        
        namen = [(f.text().strip(), n.text().strip()) for f, n in zip(self.listeVornamenFelder, self.listeNachnamenFelder)]
        
        gefilterte_namen = [(vorname, nachname) for vorname, nachname in namen if vorname or nachname]
        for i, (vorname, nachname) in enumerate(gefilterte_namen):
            if (vorname and not nachname) or (not vorname and nachname):
                QMessageBox.critical(self, "Fehler", f"Zeile {i+1}: Vor- und Nachname müssen beide angegeben werden!")
                return
            if vorname and not re.match(r'^[a-zA-ZäöüÄÖÜß\-\s]+$', vorname):
                QMessageBox.critical(self, "Fehler", f"Zeile {i+1}: Ungültiger Vorname: {vorname}")
                return
            if nachname and not re.match(r'^[a-zA-ZäöüÄÖÜß\-\s]+$', nachname):
                QMessageBox.critical(self, "Fehler", f"Zeile {i+1}: Ungültiger Nachname: {nachname}")
                return
        
        self.spielerliste = []
        for i, (vorname, nachname) in enumerate(namen):
            if vorname and nachname:
                s = Spieler(vorname, nachname, i + 1)
                self.spielerliste.append(s)
        if not self.spielerliste:
            QMessageBox.critical(self, "Fehler", "Mindestens ein Spieler muss eingetragen werden!")
            return
        
        # Starte SpielWindow
        if self.combo_modus.currentText() == "Highscore":
            self.modus = "h"
        else:
            self.modus ="r"
        
        self.spiel_window = SpielGUI(self.spielerliste, self.modus, self.ent_highscoreRunden.text().strip(), self.ent_passwort.text())
        self.spiel_window.showMaximized()
        self.close()

class SpielGUI(QMainWindow):
    def __init__(self, spielerliste, modus, hr, passwort):
        super().__init__()
        self.spielerliste = spielerliste
        self.modus = modus
        self.hr = hr
        self.passwort = passwort
        self.setWindowTitle("Dartturnier")
        self.temp_würfe = []
        self.wurf_count = 0
        self.index_spieler = 0
        self.index_runde = 1
        self.offset_runde = 0
        self.punkte = []
        self.fortschritt = 0
        self.initUI()
        
    def is_android(self):
        import platform
        return platform.system().lower() == 'android' or 'android' in sys.platform.lower()

    def initUI(self):
        central_widget = QWidget()                                                                  # Erstelle das zentrale Widget, das alle GUI-Elemente enthält
        self.setCentralWidget(central_widget)                                                       # Setze das zentrale Widget als Hauptinhalt des Fensters
        main_layout = QHBoxLayout(central_widget)                                                   # Erstelle ein horizontales Layout, um das Fenster in linke (Dartscheibe) und rechte (Tabelle) Hälfte zu teilen
    
        # Definiere Stylesheets für die GUI, abhängig von der Plattform (Android oder Desktop)
        # Stylesheets steuern das Erscheinungsbild von Widgets (z. B. Schriftgröße, Abstände)
        if self.is_android():                                                                       # Für Android: Größere Schriftgrößen und Abstände für Touch-Bedienung
            self.setStyleSheet("""
                QLabel { font-size: 16pt; }
                QTableWidget { font-size: 16pt; }
                QHeaderView::section { font-size: 16pt; padding: 8px; }
            """)
        else:                                                                                       # Für Desktop (z. B. Linux, Windows): Kleinere Schriftgrößen für Standardmonitore
            self.setStyleSheet("""                                                                 
                QLabel { font-size: 12pt; }
                QTableWidget { font-size: 12pt; }
                QHeaderView::section { font-size: 12pt; padding: 4px; }
            """)
    
        # Linke Seite: Dartscheibe
        dartscheibe_layout = QVBoxLayout()                                                          # Erstelle ein vertikales Layout für die Dartscheibe, um sie zentriert darzustellen
        self.dartscheibe_label = DartscheibeLabel()                                                 # Erstelle ein QLabel, das die Dartscheibe-Grafik anzeigen wird
        self.dartscheibe_label.set_spiel_gui(self)
        try:                                                                                        # Versuche, Ressource zu laden
            pixmap = QPixmap(":/dartscheibe.png")
            if pixmap.isNull():
                # Fallback: Lade aus Dateisystem
                pixmap = QPixmap("dartscheibe.png")
            if pixmap.isNull():                                                                     # Prüfe, ob das Bild erfolgreich geladen wurde
                raise FileNotFoundError("Dartscheibe-Grafik nicht gefunden")                        # Wenn das Bild nicht geladen werden konnte, löse eine Ausnahme aus
            self.dartscheibe_label.set_pixmap(pixmap)                                               # Skaliere das Bild, behalte das Seitenverhältnis bei
        except FileNotFoundError:
            self.dartscheibe_label.setText("Dartscheibe-Grafik nicht verfügbar")                    # Wenn die Grafik nicht gefunden wird, zeige eine Fehlermeldung im QLabel
            logging.error("Dartscheibe-Grafik konnte nicht geladen werden")                         # Protokolliere den Fehler in der Log-Datei (darts.log)
        self.punktzahl_label = QLabel("Punkte")
        self.status_label = QLabel("Spieler: Wähle eine Zelle, Runde: 1, Wurf: 0/3")
        dartscheibe_layout.addWidget(self.dartscheibe_label)                                        # Füge das QLabel zum vertikalen Layout hinzu
        dartscheibe_layout.addWidget(self.punktzahl_label)
        dartscheibe_layout.addWidget(self.status_label)
        
        # Fortschrittsbalken erstellen
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        dartscheibe_layout.addWidget(self.progress_bar)
        
        main_layout.addLayout(dartscheibe_layout,1)                                                   # Füge das Dartscheibe-Layout zum Hauptlayout hinzu (linke Hälfte)
    
    
    
        # Rechte Seite: Tabelle
        tabelle_layout = QVBoxLayout()                                                              # Erstelle ein vertikales Layout für die Tabelle
        self.tabelle = QTableWidget()                                                               # Erstelle eine QTableWidget für die Punktetabelle
        self.tabelle.setRowCount(len(self.spielerliste))                                            # Setze die Anzahl der Zeilen basierend auf der Anzahl der Spieler
        self.tabelle.setColumnCount(const_anzeigeRunden + 1)                                        # Setze die Anzahl der Spalten: 1 für Spielernamen + const_anzeigeRunden für Runden + 1 für Gesamtpunktzahl
        self.tabelle.setHorizontalHeaderLabels(["Spieler"] + [f"Runde {i+1}" for i in range(const_anzeigeRunden)]) # Definiere die Spaltenüberschriften: "Spieler", "Runde 1" bis "Runde 8", "Gesamt"
        #self.tabelle.setHorizontalHeaderLabels(["Spieler"] + [f"Runde {i+1}" for i in range(const_anzeigeRunden)] + ["Gesamt"])
        
        for i, spieler in enumerate(self.spielerliste):                                             # Fülle die erste Spalte mit Spielernamen (nicht editierbar)
            item = QTableWidgetItem(f"{spieler.vorname} {spieler.nachname}")                        # Erstelle ein Tabellenelement mit Vor- und Nachnamen des Spielers
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)                                        # Deaktiviere die Bearbeitung für Spielernamen
            self.tabelle.setItem(i, 0, item)                                                        # Setze das Element in der ersten Spalte der Zeile i
    
        for row in range(len(self.spielerliste)):                                                   # Initialisiere die Punktetabelle und Gesamtpunktzahlen
            for col in range(1, const_anzeigeRunden + 1):                                           # Fülle die Runden-Spalten (1 bis const_anzeigeRunden) mit editierbaren Feldern
                item = QTableWidgetItem("---")                                                      # Erstelle ein Tabellenelement mit Startwert "---"
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)                                    # Deaktiviere die Bearbeitung für Punkte
                self.tabelle.setItem(row, col, item)                                                # Setze das Element in der entsprechenden Zeile und Spalte
        
        self.tabelle.setColumnWidth = 40                                                            # Spaltenbreite
        self.tabelle.currentCellChanged.connect(self.zelle_ausgewaehlt)                                    # Verbinde das cellChanged-Signal mit der Methode
        self.tabelle.setMinimumWidth(300)                                                           # Setze eine Mindestbreite für die Tabelle, um sie lesbar zu halten
        self.tabelle.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)                               # Aktiviere vertikales Scrollen, wenn viele Spieler vorhanden sind
        self.tabelle.setCurrentCell(self.index_spieler,self.index_runde)
        tabelle_layout.addWidget(self.tabelle)                                                      # Füge die Tabelle zum Tabellen-Layout hinzu
        
        # Buttons für Rundenmanagement
        button_layout = QHBoxLayout()
        self.abgeben_button = QPushButton("Runde abgeben")
        self.abgeben_button.clicked.connect(self.abgabe)
        self.abgeben_button.setEnabled(False) 
        #self.überspringen_button = QPushButton("Überspringen")
        #self.überspringen_button.clicked.connect(self.runde_ueberspringen)
        button_layout.addWidget(self.abgeben_button)
        #button_layout.addWidget(self.überspringen_button)
        tabelle_layout.addLayout(button_layout)
        
        main_layout.addLayout(tabelle_layout,3)                                                       # Füge das Tabellen-Layout zum Hauptlayout hinzu (rechte Hälfte)

    def zelle_ausgewaehlt(self, current_row, current_column):
        if current_row >= 0 and 1 <= current_column <= const_anzeigeRunden:
            self.index_spieler = current_row
            self.index_runde = current_column
            self.temp_würfe = []
            self.wurf_count = 0
            #self.abgeben_button.setEnabled(False)
            self.update_status_label()
            logging.debug(f"Zelle ausgewählt: Spieler {self.index_spieler}, Runde {self.index_runde}")
        else:
            self.status_label.setText("Ungültige Zelle ausgewählt")
            logging.warning("Ungültige Zelle ausgewählt")

    def verarbeite_wurf(self, punktzahl):
        if self.wurf_count < 3:
            self.temp_würfe.append(punktzahl)
            self.wurf_count += 1
            summe = sum(self.temp_würfe)
            item = QTableWidgetItem(str(summe))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.tabelle.setItem(self.index_spieler, self.index_runde, item)
            self.update_status_label()
            logging.debug(f"Wurf {self.wurf_count}: {punktzahl}, Summe: {summe}")
            if self.wurf_count == 3:
                self.abgeben_button.setEnabled(True)
                self.finde_nächste_zelle(self.index_runde)
                #self.dartscheibe_label.setEnabled(False)  # Deaktiviere weitere Klicks
        else:
            logging.warning("Keine weiteren Würfe möglich oder Runde bereits abgeschlossen")
        self.abgeben_button.setEnabled(self.prüfe_abgabebereit())
            

    def finde_nächste_zelle(self, runde):                                                                  # Suche die nächste offene Zelle (nicht abgegebene Runde)
        for col in range(runde, const_anzeigeRunden + 1):                                               # alle Spalten durchlaufen
            for row in range(len(self.spielerliste)):                                               # alle Zeilen durchlaufen
                item = self.tabelle.item(row, col)
                if item.text() == "---":
                    self.aktueller_spieler_idx = row
                    self.aktuelle_runde = col
                    self.tabelle.setCurrentCell(row, col)
                    return
        # Falls keine offene Zelle, zurück zur ersten
        self.aktueller_spieler_idx = 0
        self.aktuelle_runde = 1
        self.tabelle.setCurrentCell(0, 1)

    def update_punktzahl_label(self, punktzahl):
        self.punktzahl_label.setText(f"Punktzahl: {punktzahl}")
        logging.debug(f"Punktzahl-Label aktualisiert: {punktzahl}")

    def update_status_label(self):
        spieler = self.spielerliste[self.index_spieler]
        self.status_label.setText(
            f"Spieler: {spieler.vorname} {spieler.nachname}, "
            f"Runde: {self.index_runde}, Wurf: {self.wurf_count}/3"
        )

    def AuswertenScheibe(self, x, y, label_width, label_height):
        r_bullseye =7.5
        r_bull = 17.5
        r_scheibe = 178
        r_dreifach = 100
        b_ring = 12
        
        # Berechne die Größe des angezeigten Bildes
        pixmap = self.dartscheibe_label.pixmap()
        if not pixmap or pixmap.isNull():
            return 0
        pixmap_size = min(pixmap.width(), pixmap.height())
    
        # Berechne die Position des Bildes im Label (zentriert)
        offset_x = (label_width - pixmap_size) / 2
        offset_y = (label_height - pixmap_size) / 2
    
        # Transformiere Mauskoordinaten relativ zum Bild
        x = x - offset_x
        y = y - offset_y
    
        # Normalisiere Koordinaten auf einen virtuellen 480x480-Raum
        scale = 480 / pixmap_size
        x = (pixmap_size / 2 - x) * scale
        y = (pixmap_size / 2 - y) * scale
        
        r = np.sqrt(x**2 + y**2)
        if r <= r_bullseye:
            punktzahl =  50
        elif r <= r_bull:
            punktzahl =  25
        elif r > r_scheibe:
            punktzahl = 0
        else:
            #Faktor bestimmen:
            if r >= r_dreifach and r <= r_dreifach + b_ring:
                faktor = 3
            elif r >= r_scheibe - b_ring and r <= r_scheibe:
                faktor = 2
            else: faktor = 1
            
            # Winkel in Grad berechnen:
            winkel = np.degrees(np.arctan2(y, -x))
            if winkel < 0:
                winkel += 360
        
            # Berechnung der Punktzahl basierend auf dem Winkel (vereinfacht):
            sektoren = [6, 13, 4, 18, 1, 20, 5, 12, 9, 14, 11, 8, 16, 7, 19, 3, 17, 2, 15, 10]
            sektorgröße = 360 / len(sektoren)
        
            sektor_index = (int((winkel + sektorgröße/2) // sektorgröße) % len(sektoren))
            punktzahl = faktor * sektoren[sektor_index]
        return punktzahl
    
    def prüfe_abgabebereit(self):
        for row in range(len(self.spielerliste)):                                               # alle Zeilen durchlaufen
            item = self.tabelle.item(row, 1)
            if item.text() == "---":
                return False
        return True
    
    def abgabe(self):
        if self.prüfe_abgabebereit():
            übertrag = []
            for row in range(len(self.spielerliste)):                                               # alle Zeilen durchlaufen
                item = self.tabelle.item(row, 1)
                übertrag.append(int(item.text()))
                for col in range(2, const_anzeigeRunden + 1):
                    item_links = self.tabelle.item(row, col - 1)
                    item_rechts = self.tabelle.item(row, col)
                    item_links.setText(item_rechts.text())
                item_ende = self.tabelle.item(row, const_anzeigeRunden)
                item_ende.setText("---")
            self.punkte.append(übertrag)
            self.offset_runde += 1
            self.tabelle.setHorizontalHeaderLabels(["Spieler"] + [f"Runde {self.offset_runde+i+1}" for i in range(const_anzeigeRunden)])
            #Cursor Tabelle verschieben
            if self.index_runde >= 2:
                self.index_runde -= 1
                self.tabelle.setCurrentCell(self.index_spieler, self.index_runde)
            #Abgabebereitschaft prüfen, ggf. Button deaktivieren
            self.abgeben_button.setEnabled(self.prüfe_abgabebereit())
            logging.debug(self.punkte)
            self.update_fortschritt()

    def update_fortschritt(self):
        if self.modus == "h":
            max_summe = 0
            for spieler in range(len(self.spielerliste)):
                summe = 0
                for runde in range(self.offset_runde):
                    summe += self.punkte[runde][spieler]
                if summe > max_summe:
                    max_summe = summe
            fortschritt = (max_summe / int(self.hr)) * 100
        elif self.modus == "r":
            fortschritt = (self.offset_runde / int(self.hr)) * 100
        else:
            fortschritt = 0
        self.progress_bar.setValue(int(fortschritt))
        if fortschritt >= 100:
            self.e = ende(self.passwort, self.spielerliste, self.punkte)
            self.e.show()
            self.close()
            


class ende(QMainWindow):
    def __init__(self, passwort, spielerliste, punkte):
        super().__init__()
        self.passwort = str(passwort)
        self.spielerliste = spielerliste
        self.punkte = punkte
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Spiel beendet')                                                # Fenster-Einstellungen
        self.setGeometry(300, 300, 300, 200)
        
        central_widget = QWidget()                                                          # Zentrales Widget erstellen
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()                                                              # Layout
        
        self.info_label = QLabel('Spiel beendet! Bitte geben Sie das Passwort ein:', self)  # Info-Label
        layout.addWidget(self.info_label)
        
        self.password_input = QLineEdit(self)                                               # Passwortfeld
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)
        
        self.check_button = QPushButton('Passwort prüfen', self)                            # Button
        self.check_button.clicked.connect(self.check_password)
        layout.addWidget(self.check_button)
        
        central_widget.setLayout(layout)                                                    # Layout auf zentrales Widget setzen
        
    def check_password(self):
        # Beispiel: korrektes Passwort ist "secret123"
        entered_password = self.password_input.text()
        
        if entered_password == self.passwort:
            QMessageBox.information(self, 'Erfolg', 'Passwort korrekt!')
            self.save_excel()
        else:
            QMessageBox.warning(self, 'Fehler', 'Falsches Passwort!')
            
        # Passwortfeld leeren
        self.password_input.clear()

    def save_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Ergebnisse speichern", "Turnierergebnisse.xlsx", "Excel-Dateien (*.xlsx);;Alle Dateien (*.*)"
        )
        if not file_path:
            logging.info("Kein Speicherort für Excel-Datei ausgewählt")
            return
        
        # Sicherstellen, dass die Dateiendung .xlsx vorhanden ist
        if not file_path.lower().endswith('.xlsx'):
            file_path += '.xlsx'
        
        try:
            # Berechne Gesamtpunktzahlen und beste Runde für jeden Spieler
            ergebnisse = []
            for i, spieler in enumerate(self.spielerliste):
                gesamtpunktzahl = 0
                beste_runde = 0
                for runde in self.punkte:
                    punkte = runde[i]
                    gesamtpunktzahl += punkte
                    if punkte > beste_runde:
                        beste_runde = punkte
                ergebnisse.append({
                    'spieler': spieler,
                    'gesamtpunktzahl': gesamtpunktzahl,
                    'beste_runde': beste_runde
                })
            
            # Sortiere nach Gesamtpunktzahl (absteigend), bei Gleichstand nach bester Runde
            ergebnisse.sort(key=lambda x: (x['gesamtpunktzahl'], x['beste_runde']), reverse=True)
            
            # Erstelle Excel-Datei
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Turnierergebnisse"
            
            # Kopfzeile
            headers = ["Platzierung", "Vorname", "Nachname", "Gesamtpunktzahl", "Beste Runde"]
            for col, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}1"] = header
            
            # Daten einfügen
            for row, ergebnis in enumerate(ergebnisse, 2):
                ws[f"A{row}"] = row - 1  # Platzierung
                ws[f"B{row}"] = ergebnis['spieler'].vorname
                ws[f"C{row}"] = ergebnis['spieler'].nachname
                ws[f"D{row}"] = ergebnis['gesamtpunktzahl']
                ws[f"E{row}"] = ergebnis['beste_runde']
            
            # Spaltenbreiten anpassen
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Speichere die Datei
            wb.save(file_path)
            QMessageBox.information(self, "Erfolg", f"Ergebnisse erfolgreich als {file_path} gespeichert!")
            logging.info(f"Excel-Datei erfolgreich gespeichert: {file_path}")
        except Exception as e:
            error_msg = str(e).encode('ascii', 'replace').decode('ascii')
            QMessageBox.critical(self, "Fehler", f"Fehler beim Speichern der Excel-Datei: {error_msg}")
            logging.error(f"Fehler beim Speichern der Excel-Datei: {error_msg}")

class test:
    def __init__(self):
        self.a = Spieler("Spieler","A",1)
        self.b = Spieler("Spieler","B",2)
        self.c = Spieler("Spieler","C",3)
        self.d = Spieler("Spieler","D",4)
        self.e = Spieler("Spieler","E",5)
        self.f = Spieler("Spieler","F",6)
        self.liste = [self.a, self.b, self.c, self.d, self.e, self.f]
        app = QApplication(sys.argv)
        window = SpielGUI(self.liste, "h", 360, 123)
        window.showMaximized()
        sys.exit(app.exec_())

def main():
    app = QApplication(sys.argv)
    window = SetupWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
    #test()